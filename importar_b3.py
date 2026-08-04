"""Importação dos relatórios da Área do Investidor da B3 (DESIGN.md §6).

Dois relatórios, formatos diferentes:

* **Negociação** — compras e vendas.
* **Movimentação** — proventos, eventos e transferências.

`ler()` não grava nada: devolve a conferência para a tela. `gravar()` grava o que
foi conferido. O arquivo original nunca é guardado — traz CPF.

Duas armadilhas do formato que moldam o módulo:

1. **"Transferência - Liquidação" na Movimentação é a entrega dos negócios que já
   vêm na Negociação.** Importar os dois arquivos sem descartar essas linhas
   duplica a carteira inteira.
2. **O relatório de Negociação não traz corretagem nem emolumentos.** Todo preço
   médio importado nasce sem custos, e a conferência precisa dizer isso.
"""
from __future__ import annotations

import csv
import hashlib
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

import textos

NEGOCIACAO = "NEGOCIACAO"
MOVIMENTACAO = "MOVIMENTACAO"

NOVA = "NOVA"
DUPLICADA = "DUPLICADA"
IGNORADA = "IGNORADA"
PENDENTE = "PENDENTE"
ERRO = "ERRO"

# Movimentação → tipo de lançamento do razão
TIPOS = {
    "dividendo": "DIVIDENDO",
    "juros sobre capital proprio": "JCP",
    "rendimento": "RENDIMENTO",
    "amortizacao": "AMORTIZACAO",
    "bonificacao em ativos": "BONIFICACAO",
    "fracao em ativos": "BONIFICACAO",
    "leilao de fracao": "VENDA",
}

# Renda fixa na Movimentação: o sentido vem de Entrada/Saída, não do rótulo.
# "COMPRA / VENDA" é literalmente o texto que a B3 usa para aplicação em CDB.
RF_MOVIMENTOS = ("aplicacao", "compra / venda", "resgate antecipado",
                 "resgate", "vencimento", "resgate total", "resgate parcial")

# A subscrição inteira fica pendente de lançamento manual. Ela anda em ativos
# intermediários — direito (…11 vira …12), recibo (…13) — e as linhas nomeiam o
# papel intermediário, não o que se recebe: "Direitos de Subscrição - Exercido"
# aparece sobre o MXRF12, mas quem entra na carteira é o MXRF11. Registrar como
# está enche a posição de fantasma (aconteceu numa importação real) e registra
# o custo no ativo errado. O usuário lança a subscrição quando ela se converter.
SUBSCRICAO = "subscricao"

# Descartadas de propósito, com motivo — nunca em silêncio
IGNORAR = {
    "transferencia - liquidacao":
        "entrega dos negócios que já vêm no relatório de Negociação",
    "compra": "já vem no relatório de Negociação",
    "venda": "já vem no relatório de Negociação",
    "desdobramento": "evento corporativo: lance em Eventos, com o fator",
    "grupamento": "evento corporativo: lance em Eventos, com o fator",
    "atualizacao": "sem efeito em posição ou caixa",
    "cessao de direitos": "lance à mão se houve alienação",
    "cessao de direitos - solicitada": "lance à mão se houve alienação",
    "direitos de subscricao - nao exercido": "sem efeito em posição",
    "emprestimo": "aluguel de ativos está fora do escopo da v1",
}

def _e_tesouro(produto: str) -> bool:
    """Tesouro Direto não passa pela bolsa: nada dele vem no relatório de
    Negociação, e a Movimentação é a única fonte da compra."""
    return _chave(produto).startswith("tesouro ")


EPS = 1e-9

SUFIXO_BDR = ("31", "32", "33", "34", "35", "39")
SUFIXO_ACAO = ("3", "4", "5", "6", "7", "8")


class ArquivoNaoReconhecido(Exception):
    pass


@dataclass
class Linha:
    n: int                      # número da linha no arquivo, para a conferência
    situacao: str
    tipo: str = ""
    data: str = ""
    ticker: str = ""
    instituicao: str = ""
    quantidade: float = 0.0
    preco: float = 0.0
    valor: float = 0.0
    hash: str = ""
    motivo: str = ""
    nome_ativo: str = ""
    destino: str = ""           # só em TRANSFERENCIA: a instituição que recebe


@dataclass
class Conferencia:
    arquivo: str
    relatorio: str
    linhas: list[Linha] = field(default_factory=list)
    ativos_novos: dict[str, dict] = field(default_factory=dict)
    instituicoes_novas: list[str] = field(default_factory=list)
    avisos: list[str] = field(default_factory=list)

    def por_situacao(self, situacao: str) -> list[Linha]:
        return [l for l in self.linhas if l.situacao == situacao]

    @property
    def novas(self) -> int:
        return len(self.por_situacao(NOVA))

    @property
    def duplicadas(self) -> int:
        return len(self.por_situacao(DUPLICADA))

    @property
    def erros(self) -> int:
        return len(self.por_situacao(ERRO))


# --------------------------------------------------------------------- leitura crua

_chave = textos.chave
_numero = textos.numero
_data = textos.data_iso


def _linhas_csv(caminho: Path) -> list[list]:
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            texto = caminho.read_text(encoding=encoding)
            break
        except UnicodeDecodeError:
            continue
    try:
        dialeto = csv.Sniffer().sniff(texto[:4096], delimiters=";,\t")
    except csv.Error:
        dialeto = csv.excel                      # arquivo de uma coluna só
        dialeto.delimiter = ";"
    return [linha for linha in csv.reader(texto.splitlines(), dialeto) if any(linha)]


def _linhas_xlsx(caminho: Path) -> list[list]:
    from openpyxl import load_workbook

    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        planilha = wb[wb.sheetnames[0]]
        # A planilha da B3 declara `<dimension ref="A1:A1"/>`, o que é mentira: o
        # modo read_only acredita e devolve só a coluna A, então o cabeçalho
        # chegava com um campo só e o arquivo era recusado como "não é da B3".
        planilha.reset_dimensions()
        return [list(linha) for linha in planilha.iter_rows(values_only=True)
                if any(c is not None and str(c).strip() for c in linha)]
    finally:
        wb.close()


def _tabela(caminho: Path) -> tuple[list[dict], str]:
    """Acha o cabeçalho e devolve dicionários com as colunas normalizadas.

    O cabeçalho nem sempre é a primeira linha: a B3 às vezes põe título antes."""
    cru = _linhas_xlsx(caminho) if caminho.suffix.lower() in (".xlsx", ".xlsm") \
        else _linhas_csv(caminho)
    for i, linha in enumerate(cru[:15]):
        colunas = [_chave(str(c or "")) for c in linha]
        if "instituicao" in colunas and any("data" in c for c in colunas):
            corpo = cru[i + 1:]
            # a convenção decimal é decidida uma vez, pelo arquivo inteiro: a
            # planilha da B3 escreve em formato americano e o CSV em brasileiro
            formato = textos.formato_numerico(c for l in corpo for c in l)
            return [dict(zip(colunas, valores)) for valores in corpo], formato
    raise ArquivoNaoReconhecido(
        f"{caminho.name}: nenhum cabeçalho com 'Instituição' e 'Data' nas 15 "
        f"primeiras linhas — o arquivo é da Área do Investidor da B3?")


def _exigir(tabela: list[dict], colunas: tuple[str, ...], relatorio: str) -> None:
    faltando = [c for c in colunas if c not in tabela[0]]
    if faltando:
        raise ArquivoNaoReconhecido(
            f"relatório de {relatorio} sem a(s) coluna(s): {', '.join(faltando)}")


# --------------------------------------------------------------------- normalização

def classe_provavel(ticker: str) -> tuple[str, bool]:
    """Devolve (classe sugerida, precisa confirmar).

    O sufixo 11 é ambíguo — FII, ETF e unit dividem o mesmo número — e a diferença
    muda a alíquota do IR. Nunca decidir sozinho nesse caso."""
    if ticker.upper().startswith("TESOURO-"):
        return "TESOURO", False     # código derivado do próprio nome do papel
    if _PREFIXO_RF.match(ticker.upper()) and _CODIGO_RF.fullmatch(ticker.upper()):
        # o prefixo já diz o que é: perguntar a classe de um código que começa
        # com CDB é pedir ao usuário para confirmar o óbvio. Sem isto, só os
        # códigos que NÃO terminam em dígito eram reconhecidos — `CDB726AM6KA`
        # passava e `CDB3268VM70` virava pergunta.
        return "RF", False
    if _CODIGO_RF.fullmatch(ticker.upper()) and not re.search(r"\d+$", ticker):
        return "RF", False          # código de papel de renda fixa: CDB726AWP4H
    sufixo = re.search(r"(\d+)$", ticker)
    if not sufixo:
        return "", True
    digitos = sufixo.group(1)
    if digitos in SUFIXO_BDR:
        return "BDR", False
    if digitos == "11":
        return "FII", True
    if digitos in SUFIXO_ACAO:
        return "ACAO", False
    return "", True


# prefixos que a B3 usa no código do papel de renda fixa
_PREFIXO_RF = re.compile(r"^(CDB|RDB|LCI|LCA|LC|LF|CRI|CRA|DEB)\d")
_CODIGO_RF = re.compile(r"\b([A-Z]{2,}\d[A-Z0-9]{4,})\b")


def _ticker_rf(produto: str) -> tuple[str, str]:
    """`CDB - CDB726AWP4H - BANCO X` -> (`CDB726AWP4H`, `CDB … BANCO X`).

    O separador é o mesmo da renda variável, mas aqui o **primeiro** pedaço é só
    o tipo do papel: partir igual faria todo CDB virar o ticker `CDB`, fundindo
    títulos diferentes numa posição só."""
    achado = _CODIGO_RF.search(str(produto).upper())
    nome = re.sub(r"\s+", " ", str(produto)).strip()
    return (achado.group(1) if achado else ""), nome


def _ticker(produto: str, mercado: str = "") -> tuple[str, str]:
    """'PETR4 - PETROLEO BRASILEIRO SA' -> ('PETR4', 'PETROLEO BRASILEIRO SA').

    No mercado fracionário o código vem com F no fim (PETR4F): é o mesmo ativo,
    e tratá-lo como outro quebraria o preço médio."""
    partes = re.split(r"\s+-\s+", str(produto).strip(), maxsplit=1)
    codigo = partes[0].strip().upper()
    nome = partes[1].strip() if len(partes) > 1 else ""
    if codigo.endswith("F") and re.search(r"\d", codigo) and \
            ("fracion" in _chave(mercado) or re.match(r"^[A-Z]{4}\d+F$", codigo)):
        codigo = codigo[:-1]
    return codigo, nome


def _hash(relatorio: str, campos: tuple, ocorrencia: int) -> str:
    """Hash canônico da linha, com o número da ocorrência dentro do arquivo.

    Sem a ocorrência, dois negócios idênticos no mesmo dia (mesma quantidade,
    mesmo preço) viram um só. Com ela, reimportar o mesmo arquivo continua
    idempotente e o negócio repetido de verdade entra duas vezes."""
    crua = "|".join([relatorio, *(f"{c}" for c in campos), str(ocorrencia)])
    return hashlib.sha256(crua.encode("utf-8")).hexdigest()


# --------------------------------------------------------------------- leitura

def _ler_negociacao(tabela: list[dict], conf: Conferencia, formato: str) -> None:
    _exigir(tabela, ("data do negocio", "tipo de movimentacao", "instituicao",
                     "codigo de negociacao", "quantidade", "preco"), "Negociação")
    conf.avisos.append(
        "O relatório de Negociação da B3 não traz corretagem nem emolumentos: "
        "os preços médios entram sem custos. Lance os custos à parte.")
    vistos: dict[str, int] = {}
    for i, linha in enumerate(tabela, start=2):
        movimento = _chave(str(linha.get("tipo de movimentacao") or ""))
        if movimento not in ("compra", "venda"):
            conf.linhas.append(Linha(i, IGNORADA,
                                     motivo=f"movimentação '{movimento}' fora de escopo"))
            continue
        try:
            data = _data(linha["data do negocio"])
            ticker, nome = _ticker(linha["codigo de negociacao"],
                                   str(linha.get("mercado") or ""))
            qtd = _numero(linha["quantidade"], formato)
            preco = _numero(linha["preco"], formato)
        except (ValueError, KeyError) as e:
            conf.linhas.append(Linha(i, ERRO, motivo=str(e)))
            continue
        valor = _numero(linha.get("valor"), formato) or qtd * preco
        instituicao = str(linha["instituicao"] or "").strip()
        campos = (movimento, data, ticker, instituicao, f"{qtd:.8f}", f"{valor:.2f}")
        chave = "|".join(campos)
        vistos[chave] = vistos.get(chave, -1) + 1
        conf.linhas.append(Linha(
            i, NOVA, movimento.upper(), data, ticker, instituicao, qtd, preco, valor,
            _hash(NEGOCIACAO, campos, vistos[chave]), nome_ativo=nome))


# A subscrição anda em ativos intermediários e nenhuma linha da B3 traz o valor
# pago. O que muda entre os subtipos é o que o usuário precisa fazer — e dizer
# "lance à mão" para os seis sem distinguir era mandar ele descobrir sozinho
# quais dos seis realmente exigem alguma coisa.
GUIA_SUBSCRICAO = {
    "direito de subscricao":
        "só o direito entrou na conta, nada a lançar. Se você exercer, o "
        "lançamento é o do recibo",
    "direito sobras de subscricao":
        "direito de sobras: nada a lançar por si só",
    "solicitacao de subscricao":
        "pedido registrado, ainda sem efeito em posição — nada a lançar",
    "direitos de subscricao - exercido":
        "exercício confirmado: o papel entra como RECIBO, na linha de recibo "
        "desta mesma importação. É lá que o valor pago é lançado",
    "recibo de subscricao":
        "É ESTA a linha que vira posição. Lance uma COMPRA do recibo com o "
        "valor que você pagou — a B3 não informa o preço. Quando o recibo "
        "virar a cota definitiva, registre uma CONVERSÃO em Eventos",
}


def _motivo_subscricao(movimento: str) -> str:
    for chave, guia in GUIA_SUBSCRICAO.items():
        if movimento.startswith(chave):
            return f"subscrição — {guia}"
    return ("subscrição: a linha nomeia o direito ou o recibo, não o ativo que "
            "entra na carteira — lance à mão, com o valor pago, quando ela se "
            "converter")


def _ler_subscricoes(coletadas: list[tuple[int, dict, str]], conf: Conferencia,
                     formato: str) -> None:
    """A subscrição anda em papéis intermediários, mas **o preço está lá**.

    A linha "Direitos de Subscrição - Exercido" traz preço unitário e valor da
    operação; é o que o subscritor pagou. O recibo (`…13`) é o que entra na
    carteira. Com os dois, o lançamento sai sozinho — antes o programa mandava
    lançar à mão dizendo que a B3 não informava o preço, o que **não era
    verdade**.

    Sem o par (exercício sem recibo, ou recibo sem exercício) a linha continua
    pendente: inventar o custo de uma subscrição é inventar preço médio."""
    exercidos: dict[str, dict] = {}
    for _i, linha, movimento in coletadas:
        if not movimento.startswith("direitos de subscricao - exercido"):
            continue
        preco = _numero(linha.get("preco unitario"), formato)
        valor = _numero(linha.get("valor da operacao"), formato)
        if preco or valor:
            exercidos[_data(linha["data"])] = {
                "preco": preco, "valor": valor,
                "quantidade": _numero(linha.get("quantidade"), formato),
                "instituicao": str(linha.get("instituicao") or "").strip()}

    vistos: dict[str, int] = {}
    for i, linha, movimento in coletadas:
        ticker, nome = _ticker(linha.get("produto", ""))
        data = _data(linha["data"]) if linha.get("data") else ""
        quantidade = _numero(linha.get("quantidade"), formato)
        if not movimento.startswith("recibo de subscricao"):
            # direito, sobra, solicitação e exercício não têm efeito em
            # posição: são informativos, e marcá-los como pendentes enchia a
            # conferência de linhas que não pedem nada — cinco das seis de uma
            # subscrição real
            conf.linhas.append(Linha(
                i, IGNORADA, ticker=ticker, data=data, quantidade=quantidade,
                motivo=_motivo_subscricao(movimento)))
            continue

        # o exercício é do dia anterior ou do mesmo dia: procura para trás
        pago = next((exercidos[d] for d in sorted(exercidos, reverse=True)
                     if d <= data and abs(exercidos[d]["quantidade"] - quantidade) < EPS),
                    None)
        if pago is None:
            conf.linhas.append(Linha(
                i, PENDENTE, ticker=ticker, data=data, quantidade=quantidade,
                motivo=_motivo_subscricao(movimento)))
            continue
        preco = pago["preco"] or (pago["valor"] / quantidade if quantidade else 0.0)
        valor = pago["valor"] or round(quantidade * preco, 2)
        campos = ("subscricao", data, ticker, pago["instituicao"],
                  f"{quantidade:.8f}", f"{valor:.2f}")
        chave = "|".join(campos)
        vistos[chave] = vistos.get(chave, -1) + 1
        conf.linhas.append(Linha(
            i, NOVA, "SUBSCRICAO", data, ticker, pago["instituicao"], quantidade,
            preco, valor, _hash(MOVIMENTACAO, campos, vistos[chave]),
            motivo=(f"recibo de subscrição, com o preço que a linha de exercício "
                    f"informou. Quando o recibo virar a cota definitiva, registre "
                    f"uma CONVERSÃO em Eventos"),
            nome_ativo=nome))


def _parear_transferencias(coletadas: list[tuple[int, dict]], conf: Conferencia,
                           formato: str) -> None:
    """A portabilidade vem em duas linhas: débito na origem, crédito no destino.

    Juntas, elas dizem tudo que um lançamento de transferência precisa, e o par
    vira um lançamento só. **Uma linha sozinha continua pendente**: metade da
    portabilidade não diz para onde o papel foi, e lançar assim moveria a posição
    para o nada.

    Transferência **move** posição entre corretoras, não cria. Se o papel veio de
    uma corretora que o Peculium nunca viu, o que falta é a compra original — e
    é o aviso de saldo negativo do razão que vai apontar isso."""
    lados: dict[tuple, dict[str, tuple[int, dict]]] = {}
    for i, linha in coletadas:
        ticker = _ticker(linha.get("produto", ""))[0]
        data = _data(linha["data"]) if linha.get("data") else ""
        quantidade = _numero(linha.get("quantidade"), formato)
        entrada = _chave(str(linha.get("entrada/saida") or "")).startswith("cred")
        chave = (data, ticker, round(quantidade, 8))
        lados.setdefault(chave, {})[("entrada" if entrada else "saida")] = (i, linha)

    vistos: dict[str, int] = {}
    for (data, ticker, quantidade), par in lados.items():
        saida, entrada = par.get("saida"), par.get("entrada")
        if saida and entrada:
            origem = str(saida[1].get("instituicao") or "").strip()
            destino = str(entrada[1].get("instituicao") or "").strip()
            if textos.nome_instituicao(origem) == textos.nome_instituicao(destino):
                # Débito e crédito na MESMA corretora não é portabilidade: é
                # troca de conta ou de custódia dentro dela, e não move nada
                # entre instituições. Num acervo real a B3 trouxe três dessas na
                # TORO, e cada uma virava um lançamento que a porta manual
                # RECUSA (`lancar` exige instituições diferentes) e que fazia o
                # razão gritar "saldo negativo" no débito antes de o crédito
                # entrar — três alarmes graves no painel para posição que fecha
                # em zero.
                for lado in (saida, entrada):
                    conf.linhas.append(Linha(
                        lado[0], IGNORADA, ticker=ticker, data=data,
                        instituicao=origem, quantidade=quantidade,
                        motivo=f"débito e crédito na mesma instituição "
                               f"({origem}): troca de conta ou custódia, não "
                               f"muda a posição em lugar nenhum"))
                continue
            campos = ("transferencia", data, ticker, origem, destino,
                      f"{quantidade:.8f}")
            chave = "|".join(campos)
            vistos[chave] = vistos.get(chave, -1) + 1
            conf.linhas.append(Linha(
                saida[0], NOVA, "TRANSFERENCIA", data, ticker, origem, quantidade,
                hash=_hash(MOVIMENTACAO, campos, vistos[chave]),
                destino=destino,
                motivo=f"portabilidade de {origem} para {destino}"))
            conf.linhas.append(Linha(
                entrada[0], IGNORADA,
                motivo="o outro lado da mesma portabilidade, já lançada"))
            continue
        i, linha = saida or entrada
        instituicao = str(linha.get("instituicao") or "").strip()
        conf.linhas.append(Linha(
            i, PENDENTE, ticker=ticker, data=data, instituicao=instituicao,
            quantidade=quantidade,
            motivo=(f"portabilidade {'de saída' if saida else 'de entrada'} sem o "
                    f"outro lado no arquivo: não dá para saber a corretora "
                    f"{'de destino' if saida else 'de origem'}. Lance à mão para "
                    f"preservar o custo — e note que transferência move posição, "
                    f"não cria: se o papel veio de corretora que o Peculium nunca "
                    f"viu, o que falta é a aquisição original. Se você não pagou "
                    f"nada por ele (presente, promoção), lance como BONIFICAÇÃO "
                    f"com valor zero — COMPRA recusa preço zero")))


def _ler_movimentacao(tabela: list[dict], conf: Conferencia, formato: str) -> None:
    _exigir(tabela, ("entrada/saida", "data", "movimentacao", "produto",
                     "instituicao", "quantidade"), "Movimentação")
    vistos: dict[str, int] = {}
    transferencias: list[tuple[int, dict]] = []
    subscricoes: list[tuple[int, dict, str]] = []
    for i, linha in enumerate(tabela, start=2):
        movimento = _chave(str(linha.get("movimentacao") or ""))
        # "Compra"/"Venda" na Movimentação são o outro lado do que vem na
        # Negociação — MENOS no Tesouro Direto, que não é negociado em bolsa e
        # por isso não aparece lá. Descartar essas linhas junto com as outras
        # fazia a compra do Tesouro sumir sem aviso nenhum: numa carteira real
        # eram R$ 2.079,13 de patrimônio evaporando em silêncio.
        if movimento in IGNORAR and not (
                movimento in ("compra", "venda")
                and _e_tesouro(str(linha.get("produto") or ""))):
            conf.linhas.append(Linha(i, IGNORADA, motivo=IGNORAR[movimento]))
            continue
        if movimento.startswith("transferencia"):
            # A portabilidade vem em DUAS linhas — débito na origem, crédito no
            # destino — e nenhuma delas sozinha diz para onde o papel foi. Aqui
            # só se guarda; o pareamento acontece depois, com o arquivo inteiro
            # lido, e é ele que decide entre lançar e deixar pendente.
            transferencias.append((i, linha))
            continue
        if SUBSCRICAO in movimento:
            subscricoes.append((i, linha, movimento))
            continue

        entrada = _chave(str(linha.get("entrada/saida") or "")).startswith("cred")
        renda_fixa = any(movimento.startswith(m) for m in RF_MOVIMENTOS)
        tesouro = movimento in ("compra", "venda")
        tipo = ("COMPRA" if entrada else "VENDA") if (renda_fixa or tesouro)             else TIPOS.get(movimento)
        if tipo is None:
            conf.linhas.append(Linha(i, ERRO, motivo=f"movimentação desconhecida: "
                                                     f"{linha.get('movimentacao')!r}"))
            continue
        try:
            data = _data(linha["data"])
            if tesouro:
                # mesmo código derivado do leitor de posição: sem isto o papel
                # comprado e o papel do retrato seriam dois ativos diferentes
                import importar_posicao
                nome = str(linha["produto"]).strip()
                ticker = importar_posicao._ticker_tesouro(nome)
            else:
                ticker, nome = (_ticker_rf(linha["produto"]) if renda_fixa
                                else _ticker(linha["produto"]))
            qtd = _numero(linha["quantidade"], formato)
        except (ValueError, KeyError) as e:
            conf.linhas.append(Linha(i, ERRO, motivo=str(e)))
            continue
        if (renda_fixa or tesouro) and not ticker:
            conf.linhas.append(Linha(
                i, ERRO, motivo=f"título de renda fixa sem código identificável em "
                                f"{linha.get('produto')!r}"))
            continue
        preco = _numero(linha.get("preco unitario"), formato)
        valor = _numero(linha.get("valor da operacao"), formato) or qtd * preco
        instituicao = str(linha["instituicao"] or "").strip()
        campos = (movimento, data, ticker, instituicao, f"{qtd:.8f}", f"{valor:.2f}")
        chave = "|".join(campos)
        vistos[chave] = vistos.get(chave, -1) + 1
        conf.linhas.append(Linha(
            i, NOVA, tipo, data, ticker, instituicao, qtd, preco, valor,
            _hash(MOVIMENTACAO, campos, vistos[chave]), nome_ativo=nome))

    # depois do arquivo inteiro lido: só aí se sabe se as duas pontas de cada
    # portabilidade estão presentes, e se a subscrição tem preço em alguma linha
    _parear_transferencias(transferencias, conf, formato)
    _ler_subscricoes(subscricoes, conf, formato)
    conf.linhas.sort(key=lambda l: l.n)


def ler(caminho: str | Path, conn) -> Conferencia:
    """Lê e confere. **Não grava nada** — quem grava é `gravar()`."""
    alvo = Path(caminho)
    tabela, formato = _tabela(alvo)
    if not tabela:
        raise ArquivoNaoReconhecido(f"{alvo.name}: cabeçalho sem nenhuma linha abaixo")
    relatorio = NEGOCIACAO if "codigo de negociacao" in tabela[0] else MOVIMENTACAO
    conf = Conferencia(alvo.name, relatorio)
    (_ler_negociacao if relatorio == NEGOCIACAO
     else _ler_movimentacao)(tabela, conf, formato)

    ja_no_banco = {r[0] for r in conn.execute(
        "SELECT hash_origem FROM lancamentos WHERE hash_origem IS NOT NULL")}
    conhecidos = {r[0].upper() for r in conn.execute("SELECT ticker FROM ativos")}
    instituicoes = {textos.nome_instituicao(r["nome"])
                    for r in conn.execute("SELECT nome FROM instituicoes")}
    for l in conf.linhas:
        if l.situacao != NOVA:
            continue
        if l.hash in ja_no_banco:
            l.situacao = DUPLICADA
            l.motivo = "já importada"
            continue
        if l.ticker and l.ticker not in conhecidos and l.ticker not in conf.ativos_novos:
            classe, confirmar = classe_provavel(l.ticker)
            conf.ativos_novos[l.ticker] = {"nome": l.nome_ativo, "classe": classe,
                                           "confirmar": confirmar}
        for nome_inst in (l.instituicao, l.destino):
            # a mesma corretora chega com grafias diferentes no MESMO
            # arquivo: comparar pelo texto cru criava um cadastro por
            # grafia — quatro da XP num extrato real
            chave = textos.nome_instituicao(nome_inst) if nome_inst else ""
            if chave and chave not in instituicoes and chave not in {
                    textos.nome_instituicao(n)
                    for n in conf.instituicoes_novas}:
                conf.instituicoes_novas.append(nome_inst)
    if any(a["confirmar"] for a in conf.ativos_novos.values()):
        conf.avisos.append(
            "Ticker terminado em 11 pode ser FII, ETF ou unit, e a classe muda a "
            "alíquota do IR — confirme antes de gravar.")
    return conf


# --------------------------------------------------------------------- gravação

def gravar(conn, conf: Conferencia, classes: dict[str, str] | None = None) -> int:
    """Grava as linhas NOVAS conferidas. Devolve quantas entraram.

    `classes` sobrescreve a classe sugerida por ticker — é por onde a tela
    devolve o que o usuário confirmou. **Só vale onde a confirmação foi
    pedida**: um código que começa com CDB é renda fixa e ponto, e aceitar que a
    tela diga o contrário já custou caro — uma lista de opções incompleta na
    interface fez todo CDB e o Tesouro entrarem como ação, e a reconciliação da
    renda fixa, que filtra por classe, parou de achar o lançamento da B3 e
    duplicou os aportes. Um defeito de tela não pode corromper o razão."""
    classes = {k.upper(): v for k, v in (classes or {}).items()}
    classes = {t: v for t, v in classes.items()
               if conf.ativos_novos.get(t, {}).get("confirmar", True)}
    faltando = [t for t, a in conf.ativos_novos.items()
                if not (classes.get(t) or a["classe"])]
    if faltando:
        raise ValueError(f"classe não definida para: {', '.join(sorted(faltando))}")

    agora = datetime.now(timezone.utc).isoformat(timespec="seconds")
    cur = conn.execute(
        "INSERT INTO importacoes (arquivo, tipo, em, linhas, novas, duplicadas, erros)"
        " VALUES (?,?,?,?,?,?,?)",
        (conf.arquivo, conf.relatorio, agora, len(conf.linhas), conf.novas,
         conf.duplicadas, conf.erros))
    importacao_id = cur.lastrowid

    for ticker, dados in conf.ativos_novos.items():
        conn.execute("INSERT OR IGNORE INTO ativos (ticker, nome, classe) VALUES (?,?,?)",
                     (ticker, dados["nome"] or None, classes.get(ticker) or dados["classe"]))
    import lancamentos

    for nome in conf.instituicoes_novas:
        lancamentos.instituicao(conn, nome)

    ativos = {r[0].upper(): r[1] for r in conn.execute("SELECT ticker, id FROM ativos")}
    instituicoes = {textos.nome_instituicao(r["nome"]): r["id"]
                    for r in conn.execute("SELECT nome, id FROM instituicoes")}
    origem = f"B3_{conf.relatorio}"
    gravadas = 0
    for l in conf.por_situacao(NOVA):
        cur = conn.execute(
            "INSERT OR IGNORE INTO lancamentos (data, tipo, ativo_id, instituicao_id,"
            " instituicao_destino_id, quantidade, preco, valor, origem, hash_origem,"
            " importacao_id, criado_em) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (l.data, l.tipo, ativos.get(l.ticker),
             instituicoes.get(textos.nome_instituicao(l.instituicao)),
             instituicoes.get(textos.nome_instituicao(l.destino))
             if l.destino else None,
             l.quantidade, l.preco, l.valor, origem, l.hash, importacao_id, agora))
        gravadas += cur.rowcount
    return gravadas
