"""Relatórios de **posição** da B3 (DESIGN.md §6.4).

São fotografias de um instante, não extratos de transação. Por isso este módulo
**nunca cria lançamento**: um retrato traz quantidade e valor de mercado, mas não
o custo de aquisição, e inventá-lo corromperia o preço médio e, atrás dele, o
imposto.

O que ele faz com o arquivo:

1. **Confere** a carteira calculada contra a da B3, ativo a ativo. É auditoria
   independente do sistema inteiro — numa carteira real apontou exatamente os
   três lançamentos que faltavam.
2. **Grava cotações** oficiais, sem depender de rede: preço de fechamento na
   renda variável, preço na curva na renda fixa e o valor atualizado do Tesouro
   — que é o único jeito de precificar o **Tesouro IPCA+**, cuja curva não se
   reconstrói sem o VNA oficial.
3. **Preenche o cadastro dos títulos de renda fixa**: emissor, indexador,
   emissão e vencimento vêm prontos.

Aceita os três arquivos com essa estrutura: `posicao`, `relatorio-consolidado-anual`
e `relatorio-consolidado-mensal` — mudam o nome das abas e a data do retrato, não
o conteúdo.
"""
from __future__ import annotations

import calendar
import re
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

import cotacoes
import razao
import renda_fixa
import textos

ORIGEM = "B3"

CONFERE = "CONFERE"
SO_NA_B3 = "SO_NA_B3"
SO_NO_PECULIUM = "SO_NO_PECULIUM"
QUANTIDADE_DIFERE = "QUANTIDADE_DIFERE"

# nome da aba (normalizado) -> classe do ativo. O consolidado prefixa com
# "Posição - "; o relatório de posição não.
ABAS = {
    "acoes": "ACAO", "posicao - acoes": "ACAO",
    "bdr": "BDR", "posicao - bdr": "BDR",
    "fundo de investimento": "FII", "posicao - fundos": "FII",
    "fundos": "FII", "posicao - fundo de investimento": "FII",
    "renda fixa": "RF", "posicao - renda fixa": "RF",
    "tesouro direto": "TESOURO", "posicao - tesouro direto": "TESOURO",
}

MESES = ("janeiro", "fevereiro", "marco", "abril", "maio", "junho", "julho",
         "agosto", "setembro", "outubro", "novembro", "dezembro")


class ArquivoNaoReconhecido(Exception):
    pass


@dataclass
class Item:
    ticker: str
    nome: str
    classe: str
    quantidade: float
    preco: float | None
    valor: float | None
    instituicao: str = ""
    cnpj: str = ""
    emissor: str = ""
    indexador: str = ""
    emissao: str | None = None
    vencimento: str | None = None


@dataclass
class Divergencia:
    ticker: str
    situacao: str
    no_peculium: float
    na_b3: float
    classe: str = ""
    observacao: str = ""


@dataclass
class Conferencia:
    arquivo: str
    data: str
    itens: list[Item] = field(default_factory=list)
    divergencias: list[Divergencia] = field(default_factory=list)
    avisos: list[str] = field(default_factory=list)

    @property
    def confere(self) -> int:
        return sum(1 for d in self.divergencias if d.situacao == CONFERE)

    @property
    def problemas(self) -> list[Divergencia]:
        return [d for d in self.divergencias if d.situacao != CONFERE]


# --------------------------------------------------------------------- leitura

def data_de_referencia(caminho: Path) -> tuple[str, str]:
    """A data do retrato, tirada do nome do arquivo. Devolve (data, aviso).

    Importa acertar: gravar o preço do consolidado de 2025 como se fosse de hoje
    reescreveria a cotação atual com um preço de um ano atrás."""
    nome = textos.chave(caminho.stem)
    achado = re.search(r"(\d{4})-(\d{2})-(\d{2})", nome)
    if achado:
        return "-".join(achado.groups()), ""
    achado = re.search(r"anual-(\d{4})", nome)
    if achado:
        return f"{achado.group(1)}-12-31", ""
    achado = re.search(r"mensal-(\d{4})-([a-z]+)", nome)
    if achado and achado.group(2) in MESES:
        ano, mes = int(achado.group(1)), MESES.index(achado.group(2)) + 1
        return f"{ano:04d}-{mes:02d}-{calendar.monthrange(ano, mes)[1]:02d}", ""
    hoje = date.today().isoformat()
    return hoje, (f"o nome do arquivo não diz a data do retrato; assumindo hoje "
                  f"({textos.data_br(hoje)}). Se for de outra data, os preços "
                  f"entram na data errada")


def _abas(caminho: Path) -> list[tuple[str, list[dict]]]:
    from openpyxl import load_workbook

    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        saida = []
        for nome in wb.sheetnames:
            classe = ABAS.get(textos.chave(nome))
            if classe is None:
                continue
            ws = wb[nome]
            ws.reset_dimensions()      # a B3 declara dimensão falsa
            linhas = [list(l) for l in ws.iter_rows(values_only=True)
                      if any(c is not None and str(c).strip() for c in l)]
            if len(linhas) < 2:
                continue
            colunas = [textos.chave(c) for c in linhas[0]]
            saida.append((classe, [dict(zip(colunas, v)) for v in linhas[1:]]))
        return saida
    finally:
        wb.close()


def _texto(linha: dict, *nomes: str) -> str:
    for nome in nomes:
        valor = linha.get(nome)
        if valor not in (None, "", "-"):
            return str(valor).strip()
    return ""


def _ticker_tesouro(nome: str) -> str:
    """O Tesouro não tem código de negociação: `Tesouro IPCA+ com Juros
    Semestrais 2037` vira `TESOURO-IPCA-2037`, que é estável e distingue os
    papéis pelo par indexador/vencimento."""
    limpo = textos.chave(nome).upper()
    indexador = ("IPCA" if "IPCA" in limpo else
                 "SELIC" if "SELIC" in limpo else
                 "PREFIXADO" if "PREFIXADO" in limpo else "TESOURO")
    juros = "-JUROS" if "JUROS SEMESTRAIS" in limpo else ""
    ano = re.search(r"(20\d{2})", limpo)
    return f"TESOURO-{indexador}{juros}-{ano.group(1) if ano else 'S-DATA'}"


def _item(classe: str, linha: dict) -> Item | None:
    quantidade = textos.numero(linha.get("quantidade"), textos.US)
    if quantidade <= 0:
        return None                       # linha de total, ou papel zerado
    nome = _texto(linha, "produto")
    if classe == "TESOURO":
        valor = textos.numero(_texto(linha, "valor atualizado", "valor bruto"),
                              textos.US)
        return Item(_ticker_tesouro(nome), nome, classe, quantidade,
                    (valor / quantidade) if valor else None, valor or None,
                    instituicao=_texto(linha, "instituicao"),
                    indexador=_texto(linha, "indexador"),
                    vencimento=_data(linha, "vencimento"))
    if classe == "RF":
        preco = textos.numero(
            _texto(linha, "preco atualizado curva", "preco atualizado mtm",
                   "preco atualizado fechamento"), textos.US)
        valor = textos.numero(
            _texto(linha, "valor atualizado curva", "valor atualizado mtm",
                   "valor atualizado fechamento"), textos.US)
        return Item(_texto(linha, "codigo"), nome, classe, quantidade,
                    preco or None, valor or None,
                    instituicao=_texto(linha, "instituicao"),
                    emissor=_texto(linha, "emissor"),
                    indexador=_texto(linha, "indexador"),
                    emissao=_data(linha, "data de emissao"),
                    vencimento=_data(linha, "vencimento"))
    preco = textos.numero(_texto(linha, "preco de fechamento"), textos.US)
    return Item(_texto(linha, "codigo de negociacao"), nome, classe, quantidade,
                preco or None,
                textos.numero(_texto(linha, "valor atualizado"), textos.US) or None,
                instituicao=_texto(linha, "instituicao"),
                cnpj=re.sub(r"\D", "", _texto(linha, "cnpj da empresa", "cnpj do fundo")))


def _data(linha: dict, coluna: str) -> str | None:
    bruto = _texto(linha, coluna)
    try:
        return textos.data_iso(bruto) if bruto else None
    except ValueError:
        return None


def ler(caminho: str | Path) -> Conferencia:
    alvo = Path(caminho)
    abas = _abas(alvo)
    if not abas:
        raise ArquivoNaoReconhecido(
            f"{alvo.name}: nenhuma aba de posição reconhecida. Este leitor espera "
            f"o relatório de Posição ou o consolidado (anual ou mensal) da Área do "
            f"Investidor da B3")
    referencia, aviso = data_de_referencia(alvo)
    conf = Conferencia(alvo.name, referencia, avisos=[aviso] if aviso else [])
    for classe, linhas in abas:
        for linha in linhas:
            item = _item(classe, linha)
            if item and item.ticker:
                conf.itens.append(item)
    return conf


# --------------------------------------------------------------------- conferência

def conferir(conn, conf: Conferencia) -> Conferencia:
    """Compara a carteira calculada com a que a B3 informa.

    A comparação é na **data do retrato**, não hoje: senão um consolidado antigo
    apareceria divergindo de tudo o que foi comprado depois dele."""
    meu = {p.ticker.upper(): p.quantidade for p in razao.carteira(conn, conf.data)}
    deles = {i.ticker.upper(): i for i in conf.itens}

    for ticker in sorted(set(meu) | set(deles)):
        aqui = meu.get(ticker, 0.0)
        item = deles.get(ticker)
        la = item.quantidade if item else 0.0
        if abs(aqui - la) < 1e-9:
            situacao, obs = CONFERE, ""
        elif not item:
            situacao = SO_NO_PECULIUM
            obs = ("está na sua carteira e não na da B3 nesta data — confira se "
                   "houve venda ou transferência que não foi lançada")
        elif aqui == 0:
            situacao = SO_NA_B3
            obs = ("a B3 tem e o Peculium não — falta o lançamento de compra, "
                   "provavelmente anterior ao período que você importou")
        else:
            situacao = QUANTIDADE_DIFERE
            obs = "a quantidade não bate: falta lançamento, ou algum entrou dobrado"
        conf.divergencias.append(Divergencia(
            ticker, situacao, aqui, la, item.classe if item else "", obs))
    return conf


# --------------------------------------------------------------------- gravação

def gravar(conn, conf: Conferencia) -> dict:
    """Grava **cotações e cadastro**, nunca lançamento.

    Retrato não tem custo de aquisição: criar posição a partir dele inventaria
    preço médio e contaminaria o imposto. O que falta, o usuário lança."""
    ativos = {str(r[0]).upper(): r[1] for r in conn.execute("SELECT ticker, id FROM ativos")}
    novos = cotadas = titulos = 0

    for item in conf.itens:
        ativo_id = ativos.get(item.ticker.upper())
        if ativo_id is None:
            ativo_id = conn.execute(
                "INSERT INTO ativos (ticker, nome, classe, cnpj) VALUES (?,?,?,?)",
                (item.ticker.upper(), item.nome or None, item.classe,
                 item.cnpj or None)).lastrowid
            ativos[item.ticker.upper()] = ativo_id
            novos += 1
        elif item.cnpj:
            conn.execute("UPDATE ativos SET cnpj=coalesce(cnpj, ?) WHERE id=?",
                         (item.cnpj, ativo_id))

        if item.preco:
            # origem B3 não vence preço digitado à mão, igual à cotação online
            if cotacoes.registrar(conn, ativo_id, conf.data, item.preco, ORIGEM):
                cotadas += 1

        if item.classe in ("RF", "TESOURO") and item.emissao and \
                renda_fixa.titulo(conn, ativo_id) is None:
            indexador = _indexador(item.indexador)
            if not indexador:
                # acontece de verdade: a B3 deixa o indexador em branco em boa
                # parte dos CDBs. Sem ele não dá para cadastrar o título — mas o
                # preço da própria B3 já foi gravado, então a posição fica certa.
                conf.avisos.append(
                    f"{item.ticker}: a B3 não informou o indexador, então o "
                    f"título não foi cadastrado. O preço dela foi gravado e a "
                    f"posição fica correta; para ter a curva, cadastre o "
                    f"título ou importe a nota de renda fixa")
                continue
            try:
                renda_fixa.cadastrar(
                    conn, ativo_id=ativo_id, emissao=item.emissao,
                    indexador=indexador, taxa=0.0,
                    pu_base=_pu_de_emissao(conn, ativo_id, item.emissao),
                    vencimento=item.vencimento, emissor=item.emissor,
                    obs="da posição da B3 — taxa não informada")
            except ValueError as e:
                conf.avisos.append(f"{item.ticker}: {e}")
                continue
            titulos += 1
    return {"ativos_novos": novos, "cotacoes": cotadas, "titulos": titulos}


def _pu_de_emissao(conn, ativo_id: int, emissao: str) -> float:
    """O PU de emissão vem da aplicação, nunca do retrato.

    O relatório traz o preço de **hoje**: um CDB emitido a R$ 1,00 e valendo
    R$ 1,0295 entraria com a base 3% alta e toda a curva subiria junto. Quando
    não há aplicação lançada na data da emissão, R$ 1,00 é a convenção da B3."""
    linha = conn.execute(
        "SELECT preco FROM lancamentos WHERE ativo_id=? AND tipo='COMPRA'"
        "  AND data=? AND preco > 0 AND estorna_id IS NULL"
        " ORDER BY id LIMIT 1", (ativo_id, emissao)).fetchone()
    return float(linha[0]) if linha else 1.0


def _indexador(bruto: str) -> str:
    t = (bruto or "").strip().upper()
    if "IPCA" in t:
        return "IPCA"
    if "CDI" in t or "DI" == t:
        return "CDI"
    if "PRE" in t:
        return "PRE"
    return ""
